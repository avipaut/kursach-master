from flask import Blueprint, render_template, request, redirect, url_for, send_file, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import login_required, current_user
from fpdf import FPDF
from fpdf.enums import XPos, YPos
import os
import pandas as pd
from openpyxl import Workbook
import json
from datetime import datetime, timedelta
import calendar
from sqlalchemy import func
from routes.models import db, Board, List, Card, Todo, User, PriorityLevel

reports_bp = Blueprint('reports', __name__, url_prefix='/reports')
EXPORT_FOLDER = 'exports'
os.makedirs(EXPORT_FOLDER, exist_ok=True)

# === Report Types ===
REPORT_TYPES = {
    'project_status': 'Project Status',
    'user_workload': 'User Workload',
    'deadline_statistics': 'Deadline Statistics',
    'completion_rates': 'Completion Rates',
    'priority_distribution': 'Priority Distribution'
}

# === Helper Functions ===
def get_date_range(period_type):
    """Generate date range based on period type"""
    today = datetime.now().date()
    
    if period_type == 'today':
        return today, today
    elif period_type == 'yesterday':
        yesterday = today - timedelta(days=1)
        return yesterday, yesterday
    elif period_type == 'this_week':
        start_of_week = today - timedelta(days=today.weekday())
        return start_of_week, today
    elif period_type == 'last_week':
        start_of_last_week = today - timedelta(days=today.weekday() + 7)
        end_of_last_week = start_of_last_week + timedelta(days=6)
        return start_of_last_week, end_of_last_week
    elif period_type == 'this_month':
        start_of_month = today.replace(day=1)
        return start_of_month, today
    elif period_type == 'last_month':
        # First day of current month
        first_day_current = today.replace(day=1)
        # Last day of previous month
        last_day_previous = first_day_current - timedelta(days=1)
        # First day of previous month
        first_day_previous = last_day_previous.replace(day=1)
        return first_day_previous, last_day_previous
    elif period_type == 'custom':
        # Handled separately when custom dates are provided
        return None, None
    else:
        # Default to last 30 days
        return today - timedelta(days=30), today

# === Data Collection Functions ===
def get_project_status_data(board_id=None, period_start=None, period_end=None):
    """Get project status data for reports"""
    query = Card.query
    
    if board_id:
        # Find all lists in the specified board
        lists = List.query.filter_by(board_id=board_id).all()
        list_ids = [list_obj.id for list_obj in lists]
        query = query.filter(Card.list_id.in_(list_ids))
    
    # Apply date filters if provided
    if period_start:
        query = query.filter(Card.created_at >= period_start)
    if period_end:
        # Include the entire end day
        end_date = datetime.combine(period_end, datetime.max.time())
        query = query.filter(Card.created_at <= end_date)
    
    # Get all cards matching the criteria
    cards = query.all()
    
    # Calculate statistics
    total_cards = len(cards)
    completed_cards = sum(1 for card in cards if card.completed)
    completion_rate = (completed_cards / total_cards * 100) if total_cards > 0 else 0
    
    # Cards by list
    cards_by_list = {}
    for card in cards:
        list_obj = List.query.get(card.list_id)
        if list_obj:
            list_name = list_obj.name
            if list_name not in cards_by_list:
                cards_by_list[list_name] = {
                    'total': 0,
                    'completed': 0,
                    'completion_rate': 0
                }
            cards_by_list[list_name]['total'] += 1
            if card.completed:
                cards_by_list[list_name]['completed'] += 1
    
    # Calculate completion rate for each list
    for list_name in cards_by_list:
        list_data = cards_by_list[list_name]
        list_data['completion_rate'] = (list_data['completed'] / list_data['total'] * 100) if list_data['total'] > 0 else 0
    
    return {
        'total_cards': total_cards,
        'completed_cards': completed_cards,
        'completion_rate': completion_rate,
        'cards_by_list': cards_by_list
    }

def get_user_workload_data(board_id=None, period_start=None, period_end=None):
    """Get user workload data for reports"""
    query = Card.query
    
    if board_id:
        # Find all lists in the specified board
        lists = List.query.filter_by(board_id=board_id).all()
        list_ids = [list_obj.id for list_obj in lists]
        query = query.filter(Card.list_id.in_(list_ids))
    
    # Apply date filters if provided
    if period_start:
        query = query.filter(Card.created_at >= period_start)
    if period_end:
        # Include the entire end day
        end_date = datetime.combine(period_end, datetime.max.time())
        query = query.filter(Card.created_at <= end_date)
    
    # Get all cards matching the criteria
    cards = query.all()
    
    # Calculate workload by user
    user_workload = {}
    for card in cards:
        if card.assigned_to:
            user = User.query.get(card.assigned_to)
            if user:
                username = user.username
                if username not in user_workload:
                    user_workload[username] = {
                        'total': 0,
                        'completed': 0,
                        'overdue': 0,
                        'priority': {
                            'low': 0,
                            'medium': 0,
                            'high': 0,
                        }
                    }
                user_workload[username]['total'] += 1
                if card.completed:
                    user_workload[username]['completed'] += 1
                if card.deadline and card.deadline < datetime.now() and not card.completed:
                    user_workload[username]['overdue'] += 1
                
                # Count by priority - исправленная обработка приоритетов
                if hasattr(card, 'priority') and card.priority:
                    priority = card.priority.name.lower() if hasattr(card.priority, 'name') else str(card.priority).lower()
                    if priority in user_workload[username]['priority']:
                        user_workload[username]['priority'][priority] += 1
    
    return {
        'user_workload': user_workload
    }

def get_deadline_statistics(board_id=None, period_start=None, period_end=None):
    """Get deadline statistics for reports"""
    query = Card.query
    
    if board_id:
        # Find all lists in the specified board
        lists = List.query.filter_by(board_id=board_id).all()
        list_ids = [list_obj.id for list_obj in lists]
        query = query.filter(Card.list_id.in_(list_ids))
    
    # Apply date filters if provided
    if period_start:
        query = query.filter(Card.created_at >= period_start)
    if period_end:
        # Include the entire end day
        end_date = datetime.combine(period_end, datetime.max.time())
        query = query.filter(Card.created_at <= end_date)
    
    # Get all cards matching the criteria
    cards = query.all()
    
    # Calculate deadline statistics
    cards_with_deadline = [card for card in cards if card.deadline]
    total_with_deadline = len(cards_with_deadline)
    
    now = datetime.now()
    overdue = sum(1 for card in cards_with_deadline if card.deadline < now and not card.completed)
    completed_on_time = sum(1 for card in cards_with_deadline if card.completed and card.deadline >= now)
    completed_late = sum(1 for card in cards_with_deadline if card.completed and card.deadline < now)
    
    upcoming_deadlines = {}
    for card in cards_with_deadline:
        if not card.completed and card.deadline > now:
            days_left = (card.deadline.date() - now.date()).days
            if days_left <= 1:
                period = "today_tomorrow"
            elif days_left <= 7:
                period = "this_week"
            elif days_left <= 30:
                period = "this_month"
            else:
                period = "future"
                
            if period not in upcoming_deadlines:
                upcoming_deadlines[period] = []
                
            upcoming_deadlines[period].append({
                'id': card.id,
                'title': card.title,
                'deadline': card.deadline.strftime('%Y-%m-%d %H:%M'),
                'days_left': days_left,
                'assigned_to': User.query.get(card.assigned_to).username if card.assigned_to else None
            })
    
    return {
        'total_cards': len(cards),
        'total_with_deadline': total_with_deadline,
        'overdue': overdue,
        'completed_on_time': completed_on_time,
        'completed_late': completed_late,
        'upcoming_deadlines': upcoming_deadlines
    }

def get_priority_distribution(board_id=None, period_start=None, period_end=None):
    """Get priority distribution data for reports"""
    query = Card.query
    
    if board_id:
        lists = List.query.filter_by(board_id=board_id).all()
        list_ids = [list_obj.id for list_obj in lists]
        query = query.filter(Card.list_id.in_(list_ids))
    
    if period_start:
        query = query.filter(Card.created_at >= period_start)
    if period_end:
        end_date = datetime.combine(period_end, datetime.max.time())
        query = query.filter(Card.created_at <= end_date)
    
    cards = query.all()
    
    # Инициализация словарей на основе PriorityLevel
    priority_counts = {priority.name.lower(): 0 for priority in PriorityLevel}
    priority_completion = {
        priority.name.lower(): {'total': 0, 'completed': 0} for priority in PriorityLevel
    }
    priority_tasks = {priority.name.lower(): [] for priority in PriorityLevel}
    
    for card in cards:
        if not hasattr(card, 'priority') or not card.priority:
            continue  # Пропускаем карточки без приоритета
        priority = card.priority.name.lower()  # Извлекаем имя приоритета
        if priority in priority_counts:
            priority_counts[priority] += 1
            priority_completion[priority]['total'] += 1
            if card.completed:
                priority_completion[priority]['completed'] += 1
            
            task_data = {
                'id': card.id,
                'title': card.title,
                'completed': card.completed,
                'deadline': card.deadline.strftime('%Y-%m-%d') if card.deadline else None,
                'assigned_to': User.query.get(card.assigned_to).username if card.assigned_to else None
            }
            priority_tasks[priority].append(task_data)
    
    for priority in priority_completion:
        data = priority_completion[priority]
        data['completion_rate'] = (data['completed'] / data['total'] * 100) if data['total'] > 0 else 0
    
    return {
        'priority_counts': priority_counts,
        'priority_completion': priority_completion,
        'priority_tasks': priority_tasks
    }

# === Routes ===
@reports_bp.route('/')
@login_required
def reports_dashboard():
    """Main reports dashboard"""
    # Get all boards for the filter dropdown
    is_admin = getattr(current_user, 'is_admin', False)
    
    if is_admin:
        boards = Board.query.all()
    else:
        # Non-admins see only public boards and their own boards
        if hasattr(Board, 'admin_only'):
            boards = Board.query.filter(
                (Board.user_id == current_user.id) | (Board.admin_only == False)
            ).all()
        else:
            boards = Board.query.all()
    
    # Calculate quick stats
    total_active_cards = Card.query.filter_by(completed=False).count()
    completed_cards = Card.query.filter_by(completed=True).count()
    
    # Cards with upcoming deadlines (due in next 7 days)
    upcoming_deadline = Card.query.filter(
        Card.deadline > datetime.now(),
        Card.deadline <= (datetime.now() + timedelta(days=7)),
        Card.completed == False
    ).count()
    
    # Overdue cards
    overdue_cards = Card.query.filter(
        Card.deadline < datetime.now(),
        Card.completed == False
    ).count()
    
    return render_template(
        'reports/reports.html',
        boards=boards,
        report_types=REPORT_TYPES,
        is_admin=is_admin,
        quick_stats={
            'total_active': total_active_cards,
            'completed': completed_cards,
            'upcoming': upcoming_deadline,
            'overdue': overdue_cards
        }
    )

@reports_bp.route('/generate', methods=['GET'])
@login_required
def generate_report():
    """Generate report based on parameters"""
    report_type = request.args.get('report_type')
    board_id = request.args.get('board_id')
    period_type = request.args.get('period_type', 'this_month')
    
    if board_id:
        board_id = int(board_id)
    
    # Get date range based on period type
    start_date, end_date = get_date_range(period_type)
    
    # Handle custom date range
    if period_type == 'custom':
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid date format. Please use YYYY-MM-DD.', 'error')
                return redirect(url_for('reports.reports_dashboard'))
    
    # Get board name if board_id is provided
    board_name = "All Boards"
    if board_id:
        board = Board.query.get(board_id)
        if board:
            board_name = board.name
    
    # Generate report data based on type
    report_data = {}
    if report_type == 'project_status':
        report_data = get_project_status_data(board_id, start_date, end_date)
    elif report_type == 'user_workload':
        report_data = get_user_workload_data(board_id, start_date, end_date)
    elif report_type == 'deadline_statistics':
        report_data = get_deadline_statistics(board_id, start_date, end_date)
    elif report_type == 'priority_distribution':
        report_data = get_priority_distribution(board_id, start_date, end_date)
    elif report_type == 'completion_rates':
        # This combines project status data with additional time analysis
        project_data = get_project_status_data(board_id, start_date, end_date)
        priority_data = get_priority_distribution(board_id, start_date, end_date)
        report_data = {
            'project_status': project_data,
            'priority_data': priority_data
        }
    
    # Prepare context for the template
    context = {
        'report_type': report_type,
        'report_title': REPORT_TYPES.get(report_type, 'Custom Report'),
        'board_id': board_id,
        'board_name': board_name,
        'period_type': period_type,
        'start_date': start_date,
        'end_date': end_date,
        'report_data': report_data,
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M')
    }
    
    # Render the appropriate template based on report type
    return render_template(f'reports/{report_type}.html', **context)

@reports_bp.route('/export/<report_type>', methods=['GET'])
@login_required
def export_report(report_type):
    """Export report data in various formats"""
    export_format = request.args.get('format', 'pdf')
    board_id = request.args.get('board_id')
    period_type = request.args.get('period_type', 'this_month')
    
    if board_id:
        board_id = int(board_id)
    
    # Get date range based on period type
    start_date, end_date = get_date_range(period_type)
    
    # Handle custom date range
    if period_type == 'custom':
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                flash('Invalid date format. Please use YYYY-MM-DD.', 'error')
                return redirect(url_for('reports.reports_dashboard'))
    
    # Get board name if board_id is provided
    board_name = "All_Boards"
    if board_id:
        board = Board.query.get(board_id)
        if board:
            board_name = board.name.replace(' ', '_')
    
    # Generate report data based on type
    report_data = {}
    if report_type == 'project_status':
        report_data = get_project_status_data(board_id, start_date, end_date)
    elif report_type == 'user_workload':
        report_data = get_user_workload_data(board_id, start_date, end_date)
    elif report_type == 'deadline_statistics':
        report_data = get_deadline_statistics(board_id, start_date, end_date)
    elif report_type == 'priority_distribution':
        report_data = get_priority_distribution(board_id, start_date, end_date)
    elif report_type == 'completion_rates':
        # This combines project status data with additional time analysis
        project_data = get_project_status_data(board_id, start_date, end_date)
        priority_data = get_priority_distribution(board_id, start_date, end_date)
        report_data = {
            'project_status': project_data,
            'priority_data': priority_data
        }
    
    # Generate filename
    date_str = datetime.now().strftime('%Y%m%d')
    filename = f"{report_type}_{board_name}_{date_str}"
    
    # Export based on format
    if export_format == 'pdf':
        return export_as_pdf(report_type, report_data, filename, board_name, start_date, end_date)
    elif export_format == 'excel':
        return export_as_excel(report_type, report_data, filename, board_name, start_date, end_date)
    else:
        flash('Unsupported export format', 'error')
        return redirect(url_for('reports.reports_dashboard'))

class PDF(FPDF):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        # Use built-in fonts instead of DejaVu
        # FPDF has built-in support for standard fonts: Courier, Helvetica, Times, Symbol, ZapfDingbats
        self.set_font('helvetica', size=12)
    
    def header(self):
        self.set_font('helvetica', 'B', 12)
        self.cell(0, 10, 'Report', ln=1, align='C')  # ln=1 — переход на новую строку
    
    def footer(self):
        self.set_y(-15)
        self.set_font('helvetica', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}', align='C')

def export_as_pdf(report_type, report_data, filename, board_name, start_date, end_date):
    pdf = PDF()
    pdf.add_page()
    pdf.set_font("helvetica", size=12)
    
    def safe_cell(w, h, txt, ln=0, align=''):
        try:
            if ln:
                pdf.cell(w, h, txt, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align=align)
            else:
                pdf.cell(w, h, txt, align=align)
        except Exception as e:
            print(f"Error writing text: {e}")
            # Fallback for problematic characters
            pdf.cell(w, h, '?', align=align)
    
    # Report title
    pdf.set_font('helvetica', 'B', 16)
    safe_cell(0, 10, f"{REPORT_TYPES.get(report_type, 'Custom Report')}", ln=True, align='C')
    
    pdf.set_font('helvetica', 'I', 12)
    safe_cell(0, 10, f"Board: {board_name.replace('_', ' ')}", ln=True)
    safe_cell(0, 10, f"Period: {start_date.strftime('%Y-%m-%d')} - {end_date.strftime('%Y-%m-%d')}", ln=True)
    safe_cell(0, 10, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True)
    pdf.ln(10)
    
    # Add report content based on type
    pdf.set_font('helvetica', '', 12)
    
    if report_type == 'project_status':
        # Overall stats
        pdf.set_font('helvetica', 'B', 14)
        safe_cell(0, 10, "Project Overview", ln=True)
        pdf.set_font('helvetica', '', 12)
        safe_cell(0, 10, f"Total Cards: {report_data['total_cards']}", ln=True)
        safe_cell(0, 10, f"Completed Cards: {report_data['completed_cards']}", ln=True)
        safe_cell(0, 10, f"Completion Rate: {report_data['completion_rate']:.1f}%", ln=True)
        
        # Lists overview
        pdf.ln(5)
        pdf.set_font('helvetica', 'B', 14)
        safe_cell(0, 10, "Lists Overview", ln=True)
        
        for list_name, list_data in report_data['cards_by_list'].items():
            pdf.set_font('helvetica', 'B', 12)
            safe_cell(0, 10, f"{list_name}", ln=True)
            pdf.set_font('helvetica', '', 12)
            safe_cell(0, 10, f"Total Cards: {list_data['total']}", ln=True)
            safe_cell(0, 10, f"Completed: {list_data['completed']}", ln=True)
            safe_cell(0, 10, f"Completion Rate: {list_data['completion_rate']:.1f}%", ln=True)
            pdf.ln(5)
            
    elif report_type == 'user_workload':
        # User workload summary
        pdf.set_font('helvetica', 'B', 14)
        safe_cell(0, 10, "User Workload Summary", ln=True)
        
        for username, user_data in report_data['user_workload'].items():
            pdf.set_font('helvetica', 'B', 12)
            safe_cell(0, 10, f"User: {username}", ln=True)
            pdf.set_font('helvetica', '', 12)
            safe_cell(0, 10, f"Total Tasks: {user_data['total']}", ln=True)
            safe_cell(0, 10, f"Completed: {user_data['completed']}", ln=True)
            safe_cell(0, 10, f"Overdue: {user_data['overdue']}", ln=True)
            
            # Completion rate
            completion_rate = (user_data['completed'] / user_data['total'] * 100) if user_data['total'] > 0 else 0
            safe_cell(0, 10, f"Completion Rate: {completion_rate:.1f}%", ln=True)
            
            # Tasks by priority
            pdf.set_font('helvetica', 'I', 12)
            safe_cell(0, 10, "Tasks by Priority:", ln=True)
            for priority, count in user_data['priority'].items():
                if count > 0:
                    safe_cell(0, 10, f"{priority.capitalize()}: {count}", ln=True)
            pdf.ln(5)
    
    elif report_type == 'deadline_statistics':
        # Deadline overview
        pdf.set_font('helvetica', 'B', 14)
        safe_cell(0, 10, "Deadline Overview", ln=True)
        pdf.set_font('helvetica', '', 12)
        safe_cell(0, 10, f"Total Cards: {report_data['total_cards']}", ln=True)
        safe_cell(0, 10, f"Cards with Deadline: {report_data['total_with_deadline']}", ln=True)
        safe_cell(0, 10, f"Overdue Cards: {report_data['overdue']}", ln=True)
        safe_cell(0, 10, f"Completed On Time: {report_data['completed_on_time']}", ln=True)
        safe_cell(0, 10, f"Completed Late: {report_data['completed_late']}", ln=True)
        
        # Upcoming deadlines
        pdf.ln(5)
        pdf.set_font('helvetica', 'B', 14)
        safe_cell(0, 10, "Upcoming Deadlines", ln=True)
        
        if 'upcoming_deadlines' in report_data:
            for period, cards in report_data['upcoming_deadlines'].items():
                if cards:
                    period_name = {
                        'today_tomorrow': 'Due Today or Tomorrow',
                        'this_week': 'Due This Week',
                        'this_month': 'Due This Month',
                        'future': 'Future Deadlines'
                    }.get(period, period)
                    
                    pdf.set_font('helvetica', 'B', 12)
                    safe_cell(0, 10, period_name, ln=True)
                    pdf.set_font('helvetica', '', 12)
                    
                    for card in cards:
                        assigned_to = card['assigned_to'] or 'Unassigned'
                        safe_cell(0, 10, f"- {card['title']} (Due: {card['deadline']}, Days Left: {card['days_left']}, Assigned: {assigned_to})", ln=True)
                    pdf.ln(5)
                    
    elif report_type == 'priority_distribution':
        # Priority counts
        pdf.set_font('helvetica', 'B', 14)
        safe_cell(0, 10, "Priority Distribution", ln=True)
        pdf.set_font('helvetica', '', 12)
        
        if 'priority_counts' in report_data:
            for priority, count in report_data['priority_counts'].items():
                safe_cell(0, 10, f"{priority.capitalize()}: {count} tasks", ln=True)
            
        # Completion by priority
        pdf.ln(5)
        pdf.set_font('helvetica', 'B', 14)
        safe_cell(0, 10, "Completion by Priority", ln=True)
        
        if 'priority_completion' in report_data:
            for priority, data in report_data['priority_completion'].items():
                if data['total'] > 0:
                    pdf.set_font('helvetica', 'B', 12)
                    safe_cell(0, 10, f"{priority.capitalize()}", ln=True)
                    pdf.set_font('helvetica', '', 12)
                    safe_cell(0, 10, f"Total: {data['total']}, Completed: {data['completed']}", ln=True)
                    safe_cell(0, 10, f"Completion Rate: {data['completion_rate']:.1f}%", ln=True)
                    pdf.ln(5)
        
        # Priority tasks
        pdf.set_font('helvetica', 'B', 14)
        safe_cell(0, 10, "Tasks by Priority", ln=True)
        
        if 'priority_tasks' in report_data:
            for priority, tasks in report_data['priority_tasks'].items():
                if tasks:
                    pdf.set_font('helvetica', 'B', 12)
                    safe_cell(0, 10, f"{priority.capitalize()} Priority Tasks", ln=True)
                    pdf.set_font('helvetica', '', 12)
                    
                    for task in tasks:
                        status = "Completed" if task['completed'] else "In Progress"
                        deadline = task['deadline'] or "No deadline"
                        assigned_to = task['assigned_to'] or "Unassigned"
                        safe_cell(0, 10, f"- {task['title']} (Status: {status}, Deadline: {deadline}, Assigned: {assigned_to})", ln=True)
                    pdf.ln(5)
    
    elif report_type == 'completion_rates':
        project_data = report_data['project_status']
        priority_data = report_data['priority_data']
        
        # Overall completion
        pdf.set_font('helvetica', 'B', 14)
        safe_cell(0, 10, "Overall Completion", ln=True)
        pdf.set_font('helvetica', '', 12)
        safe_cell(0, 10, f"Total Cards: {project_data['total_cards']}", ln=True)
        safe_cell(0, 10, f"Completed Cards: {project_data['completed_cards']}", ln=True)
        safe_cell(0, 10, f"Completion Rate: {project_data['completion_rate']:.1f}%", ln=True)
        
        # Completion by list
        pdf.ln(5)
        pdf.set_font('helvetica', 'B', 14)
        safe_cell(0, 10, "Completion by List", ln=True)
        
        for list_name, list_data in project_data['cards_by_list'].items():
            pdf.set_font('helvetica', 'B', 12)
            safe_cell(0, 10, f"{list_name}", ln=True)
            pdf.set_font('helvetica', '', 12)
            safe_cell(0, 10, f"Completion Rate: {list_data['completion_rate']:.1f}%", ln=True)
            safe_cell(0, 10, f"Completed: {list_data['completed']} of {list_data['total']}", ln=True)
            pdf.ln(3)
        
        # Completion by priority
        pdf.ln(5)
        pdf.set_font('helvetica', 'B', 14)
        safe_cell(0, 10, "Completion by Priority", ln=True)
        
        if 'priority_completion' in priority_data:
            for priority, data in priority_data['priority_completion'].items():
                if data['total'] > 0:
                    pdf.set_font('helvetica', 'B', 12)
                    safe_cell(0, 10, f"{priority.capitalize()}", ln=True)
                    pdf.set_font('helvetica', '', 12)
                    safe_cell(0, 10, f"Completion Rate: {data['completion_rate']:.1f}%", ln=True)
                    safe_cell(0, 10, f"Completed: {data['completed']} of {data['total']}", ln=True)
                    pdf.ln(3)
    
    # Save the PDF
    pdf_path = os.path.join(EXPORT_FOLDER, f"{filename}.pdf")
    pdf.output(pdf_path)
    
    return send_file(pdf_path, as_attachment=True)

def export_as_excel(report_type, report_data, filename, board_name, start_date, end_date):
    """Export report as Excel with complete information and formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    
    # Define styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    subheader_font = Font(bold=True, size=11)
    subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    regular_font = Font(size=10)
    percent_format = '0.0%'
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def apply_header_style(ws, row_num, col_start=1, col_end=None):
        """Apply header style to a row"""
        if not col_end:
            col_end = ws.max_column
        
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
    
    def apply_subheader_style(ws, row_num, col_start=1, col_end=None):
        """Apply subheader style to a row"""
        if not col_end:
            col_end = ws.max_column
        
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
    
    def apply_data_style(ws, row_num, col_start=1, col_end=None):
        """Apply data style to a row"""
        if not col_end:
            col_end = ws.max_column
        
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = regular_font
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
    
    def format_percentage(ws, row, col):
        """Format a cell as percentage"""
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, str) and '%' in cell.value:
            try:
                value = float(cell.value.replace('%', '')) / 100
                cell.value = value
                cell.number_format = percent_format
            except ValueError:
                pass
    
    def auto_adjust_columns(ws):
        """Auto-adjust column widths based on content"""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            
            for cell in col:
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            ws.column_dimensions[column].width = adjusted_width
    
    # Create metadata sheet
    ws_meta = wb.active
    ws_meta.title = 'Report Info'
    
    # Add report metadata
    ws_meta.append(['Report Type', REPORT_TYPES.get(report_type, 'Custom Report')])
    ws_meta.append(['Board', board_name.replace('_', ' ')])
    ws_meta.append(['Start Date', start_date.strftime('%Y-%m-%d')])
    ws_meta.append(['End Date', end_date.strftime('%Y-%m-%d')])
    ws_meta.append(['Generated', datetime.now().strftime('%Y-%m-%d %H:%M')])
    
    # Apply styles to metadata
    apply_header_style(ws_meta, 1, 1, 2)
    for i in range(2, 6):
        apply_data_style(ws_meta, i, 1, 2)
    
    # Auto-adjust columns width
    auto_adjust_columns(ws_meta)
    
    # Create appropriate sheets based on report type
    if report_type == 'project_status':
        # Overall stats sheet
        ws_overall = wb.create_sheet('Overall Stats')
        ws_overall.append(['Metric', 'Value'])
        apply_header_style(ws_overall, 1, 1, 2)
        
        # Add data
        rows = [
            ['Total Cards', report_data['total_cards']],
            ['Completed Cards', report_data['completed_cards']],
            ['Completion Rate', f"{report_data['completion_rate']:.1f}%"]
        ]
        
        for i, row in enumerate(rows, 2):
            ws_overall.append(row)
            apply_data_style(ws_overall, i, 1, 2)
            if 'Rate' in row[0]:
                format_percentage(ws_overall, i, 2)
        
        # Lists data sheet
        ws_lists = wb.create_sheet('Lists Data')
        ws_lists.append(['List Name', 'Total Cards', 'Completed Cards', 'Completion Rate'])
        apply_header_style(ws_lists, 1)
        
        row_num = 2
        for list_name, list_data in report_data['cards_by_list'].items():
            ws_lists.append([
                list_name, 
                list_data['total'], 
                list_data['completed'], 
                f"{list_data['completion_rate']:.1f}%"
            ])
            apply_data_style(ws_lists, row_num)
            format_percentage(ws_lists, row_num, 4)
            row_num += 1
        
        auto_adjust_columns(ws_overall)
        auto_adjust_columns(ws_lists)
    
    elif report_type == 'user_workload':
        # User workload sheet
        ws_users = wb.create_sheet('User Workload')
        ws_users.append(['Username', 'Total Tasks', 'Completed', 'Completion Rate', 'Overdue Tasks'])
        apply_header_style(ws_users, 1)
        
        # Priority by user sheet
        ws_priority = wb.create_sheet('Priority by User')
        ws_priority.append(['Username', 'Priority', 'Count'])
        apply_header_style(ws_priority, 1)
        
        row_users = 2
        row_priority = 2
        
        for username, data in report_data['user_workload'].items():
            completion_rate = f"{(data['completed'] / data['total'] * 100):.1f}%" if data['total'] > 0 else '0%'
            ws_users.append([
                username,
                data['total'],
                data['completed'],
                completion_rate,
                data['overdue']
            ])
            apply_data_style(ws_users, row_users)
            format_percentage(ws_users, row_users, 4)
            row_users += 1
            
            # Add priority data
            for priority, count in data['priority'].items():
                if count > 0:
                    ws_priority.append([username, priority.capitalize(), count])
                    apply_data_style(ws_priority, row_priority)
                    row_priority += 1
        
        # Add a detailed user tasks sheet if we have task-level data
        if any('tasks' in data for username, data in report_data['user_workload'].items()):
            ws_tasks = wb.create_sheet('User Tasks Detail')
            ws_tasks.append(['Username', 'Task', 'Status', 'Priority', 'Deadline'])
            apply_header_style(ws_tasks, 1)
            # Add task-level data if available
        
        auto_adjust_columns(ws_users)
        auto_adjust_columns(ws_priority)
    
    elif report_type == 'deadline_statistics':
        # Deadline overview sheet
        ws_overview = wb.create_sheet('Deadline Overview')
        ws_overview.append(['Metric', 'Value'])
        apply_header_style(ws_overview, 1, 1, 2)
        
        # Add data
        rows = [
            ['Total Cards', report_data['total_cards']],
            ['Cards with Deadline', report_data['total_with_deadline']],
            ['Overdue Cards', report_data['overdue']],
            ['Completed On Time', report_data['completed_on_time']],
            ['Completed Late', report_data['completed_late']]
        ]
        
        for i, row in enumerate(rows, 2):
            ws_overview.append(row)
            apply_data_style(ws_overview, i, 1, 2)
        
        # Calculate and add percentages
        if report_data['total_with_deadline'] > 0:
            overdue_percent = report_data['overdue'] / report_data['total_with_deadline'] * 100
            on_time_percent = report_data['completed_on_time'] / report_data['total_with_deadline'] * 100
            late_percent = report_data['completed_late'] / report_data['total_with_deadline'] * 100
            
            ws_overview.append(['Overdue Rate', f"{overdue_percent:.1f}%"])
            ws_overview.append(['On-Time Completion Rate', f"{on_time_percent:.1f}%"])
            ws_overview.append(['Late Completion Rate', f"{late_percent:.1f}%"])
            
            for i in range(6, 9):
                apply_data_style(ws_overview, i, 1, 2)
                format_percentage(ws_overview, i, 2)
        
        # Upcoming deadlines sheet
        ws_upcoming = wb.create_sheet('Upcoming Deadlines')
        ws_upcoming.append(['Period', 'Card Title', 'Deadline', 'Days Left', 'Assigned To'])
        apply_header_style(ws_upcoming, 1)
        
        row_num = 2
        if 'upcoming_deadlines' in report_data:
            for period, cards in report_data['upcoming_deadlines'].items():
                period_name = {
                    'today_tomorrow': 'Today/Tomorrow',
                    'this_week': 'This Week',
                    'this_month': 'This Month',
                    'future': 'Future'
                }.get(period, period)
                
                if cards:  # Add a subheader for each period
                    ws_upcoming.append([f"{period_name} ({len(cards)} cards)", "", "", "", ""])
                    apply_subheader_style(ws_upcoming, row_num, 1, 5)
                    row_num += 1
                
                for card in cards:
                    ws_upcoming.append([
                        period_name,
                        card['title'],
                        card['deadline'],
                        card['days_left'],
                        card['assigned_to'] or 'Unassigned'
                    ])
                    apply_data_style(ws_upcoming, row_num)
                    row_num += 1
        
        auto_adjust_columns(ws_overview)
        auto_adjust_columns(ws_upcoming)
    
    elif report_type == 'priority_distribution':
        # Priority counts sheet
        ws_counts = wb.create_sheet('Priority Counts')
        ws_counts.append(['Priority', 'Count', 'Percentage'])
        apply_header_style(ws_counts, 1)
        
        # Calculate total for percentage
        total_priority_cards = sum(report_data['priority_counts'].values())
        
        # Add data with percentages
        row_num = 2
        for priority, count in report_data['priority_counts'].items():
            percentage = (count / total_priority_cards * 100) if total_priority_cards > 0 else 0
            ws_counts.append([
                priority.capitalize(), 
                count, 
                f"{percentage:.1f}%"
            ])
            apply_data_style(ws_counts, row_num)
            format_percentage(ws_counts, row_num, 3)
            row_num += 1
        
        # Add total row
        ws_counts.append(['Total', total_priority_cards, '100.0%'])
        apply_subheader_style(ws_counts, row_num)
        format_percentage(ws_counts, row_num, 3)
        
        # Priority completion rates sheet
        ws_completion = wb.create_sheet('Priority Completion')
        ws_completion.append(['Priority', 'Total', 'Completed', 'Completion Rate'])
        apply_header_style(ws_completion, 1)
        
        row_num = 2
        for priority, data in report_data['priority_completion'].items():
            if data['total'] > 0:
                ws_completion.append([
                    priority.capitalize(),
                    data['total'],
                    data['completed'],
                    f"{data['completion_rate']:.1f}%"
                ])
                apply_data_style(ws_completion, row_num)
                format_percentage(ws_completion, row_num, 4)
                row_num += 1
        
        # Priority tasks sheet - with groups by priority
        ws_tasks = wb.create_sheet('Priority Tasks')
        ws_tasks.append(['Priority', 'Task Title', 'Status', 'Deadline', 'Assigned To'])
        apply_header_style(ws_tasks, 1)
        
        row_num = 2
        for priority, tasks in report_data['priority_tasks'].items():
            if tasks:
                # Add a header for each priority level
                ws_tasks.append([f"{priority.capitalize()} Priority ({len(tasks)} tasks)", "", "", "", ""])
                apply_subheader_style(ws_tasks, row_num, 1, 5)
                row_num += 1
                
                for task in tasks:
                    ws_tasks.append([
                        priority.capitalize(),
                        task['title'],
                        'Completed' if task['completed'] else 'In Progress',
                        task['deadline'] or 'No deadline',
                        task['assigned_to'] or 'Unassigned'
                    ])
                    apply_data_style(ws_tasks, row_num)
                    row_num += 1
        
        auto_adjust_columns(ws_counts)
        auto_adjust_columns(ws_completion)
        auto_adjust_columns(ws_tasks)
    
    elif report_type == 'completion_rates':
        project_data = report_data['project_status']
        priority_data = report_data['priority_data']
        
        # Overall completion sheet
        ws_overall = wb.create_sheet('Overall Completion')
        ws_overall.append(['Metric', 'Value'])
        apply_header_style(ws_overall, 1, 1, 2)
        
        # Add data
        rows = [
            ['Total Cards', project_data['total_cards']],
            ['Completed Cards', project_data['completed_cards']],
            ['Overall Completion Rate', f"{project_data['completion_rate']:.1f}%"]
        ]
        
        for i, row in enumerate(rows, 2):
            ws_overall.append(row)
            apply_data_style(ws_overall, i, 1, 2)
            if 'Rate' in row[0]:
                format_percentage(ws_overall, i, 2)
        
        # Lists completion sheet
        ws_lists = wb.create_sheet('Lists Completion')
        ws_lists.append(['List Name', 'Total Cards', 'Completed Cards', 'Completion Rate'])
        apply_header_style(ws_lists, 1)
        
        row_num = 2
        for list_name, list_data in project_data['cards_by_list'].items():
            ws_lists.append([
                list_name, 
                list_data['total'], 
                list_data['completed'], 
                f"{list_data['completion_rate']:.1f}%"
            ])
            apply_data_style(ws_lists, row_num)
            format_percentage(ws_lists, row_num, 4)
            row_num += 1
        
        # Priority completion sheet
        ws_priority = wb.create_sheet('Priority Completion')
        ws_priority.append(['Priority', 'Total', 'Completed', 'Completion Rate'])
        apply_header_style(ws_priority, 1)
        
        row_num = 2
        for priority, data in priority_data['priority_completion'].items():
            if data['total'] > 0:
                ws_priority.append([
                    priority.capitalize(),
                    data['total'],
                    data['completed'],
                    f"{data['completion_rate']:.1f}%"
                ])
                apply_data_style(ws_priority, row_num)
                format_percentage(ws_priority, row_num, 4)
                row_num += 1
        
        # Add a completion trends sheet (if we had time-series data)
        # This could be filled with data if the API provided time-based completion info
        
        auto_adjust_columns(ws_overall)
        auto_adjust_columns(ws_lists)
        auto_adjust_columns(ws_priority)
    
    # Add a summary dashboard sheet as the first visible sheet
    ws_dashboard = wb.create_sheet("Dashboard", 1)
    ws_dashboard.append(['Report Summary'])
    apply_header_style(ws_dashboard, 1, 1, 1)
    
    ws_dashboard.append(['Report Type', REPORT_TYPES.get(report_type, 'Custom Report')])
    ws_dashboard.append(['Board', board_name.replace('_', ' ')])
    ws_dashboard.append(['Period', f"{start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"])
    
    for i in range(2, 5):
        apply_data_style(ws_dashboard, i, 1, 2)
    
    # Add report-specific summary data
    ws_dashboard.append([])
    ws_dashboard.append(['Key Metrics'])
    apply_subheader_style(ws_dashboard, 6, 1, 2)
    
    if report_type == 'project_status':
        ws_dashboard.append(['Total Cards', report_data['total_cards']])
        ws_dashboard.append(['Completion Rate', f"{report_data['completion_rate']:.1f}%"])
        format_percentage(ws_dashboard, 8, 2)
        
    elif report_type == 'user_workload':
        # Find user with most tasks
        most_tasks_user = max(report_data['user_workload'].items(), 
                             key=lambda x: x[1]['total'], 
                             default=(None, {'total': 0}))
        
        ws_dashboard.append(['Total Users', len(report_data['user_workload'])])
        if most_tasks_user[0]:
            ws_dashboard.append(['Most Loaded User', f"{most_tasks_user[0]} ({most_tasks_user[1]['total']} tasks)"])
        
    elif report_type == 'deadline_statistics':
        ws_dashboard.append(['Cards with Deadline', report_data['total_with_deadline']])
        ws_dashboard.append(['Overdue Cards', report_data['overdue']])
        
        if report_data['total_with_deadline'] > 0:
            overdue_percent = report_data['overdue'] / report_data['total_with_deadline'] * 100
            ws_dashboard.append(['Overdue Rate', f"{overdue_percent:.1f}%"])
            format_percentage(ws_dashboard, 9, 2)
        
    elif report_type == 'priority_distribution':
        priority_counts = report_data['priority_counts']
        ws_dashboard.append(['High Priority Tasks', priority_counts.get('high', 0)])
        ws_dashboard.append(['Medium Priority Tasks', priority_counts.get('medium', 0)])
        ws_dashboard.append(['Low Priority Tasks', priority_counts.get('low', 0)])
        
    elif report_type == 'completion_rates':
        project_data = report_data['project_status']
        ws_dashboard.append(['Overall Completion Rate', f"{project_data['completion_rate']:.1f}%"])
        format_percentage(ws_dashboard, 7, 2)
        
        # Find list with best completion rate
        best_list = max(project_data['cards_by_list'].items(), 
                        key=lambda x: x[1]['completion_rate'], 
                        default=(None, {'completion_rate': 0}))
        
        if best_list[0]:
            ws_dashboard.append(['Best Performing List', 
                               f"{best_list[0]} ({best_list[1]['completion_rate']:.1f}%)"])
            
    # Apply styling to the rest of the dashboard
    for i in range(7, ws_dashboard.max_row + 1):
        apply_data_style(ws_dashboard, i, 1, 2)
    
    # Add sheet navigation instructions
    ws_dashboard.append([])
    ws_dashboard.append(['Available Sheets:'])
    apply_subheader_style(ws_dashboard, ws_dashboard.max_row, 1, 1)
    
    for sheet in wb.sheetnames:
        if sheet != 'Dashboard':
            ws_dashboard.append([sheet])
    
    auto_adjust_columns(ws_dashboard)
    
    # Save the Excel file
    excel_path = os.path.join(EXPORT_FOLDER, f"{filename}.xlsx")
    wb.save(excel_path)
    
    return send_file(excel_path, as_attachment=True)

@reports_bp.route('/api/data/<report_type>')
@login_required
def get_report_data(report_type):
    """API endpoint to get report data for charts"""
    board_id = request.args.get('board_id')
    period_type = request.args.get('period_type', 'this_month')
    
    if board_id:
        board_id = int(board_id)
    
    # Get date range based on period type
    start_date, end_date = get_date_range(period_type)
    
    # Handle custom date range
    if period_type == 'custom':
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'error': 'Invalid date format. Please use YYYY-MM-DD.'}), 400
    
    # Generate report data based on type
    if report_type == 'project_status':
        data = get_project_status_data(board_id, start_date, end_date)
    elif report_type == 'user_workload':
        data = get_user_workload_data(board_id, start_date, end_date)
    elif report_type == 'deadline_statistics':
        data = get_deadline_statistics(board_id, start_date, end_date)
    elif report_type == 'priority_distribution':
        data = get_priority_distribution(board_id, start_date, end_date)
    elif report_type == 'completion_rates':
        # This combines project status data with additional time analysis
        project_data = get_project_status_data(board_id, start_date, end_date)
        priority_data = get_priority_distribution(board_id, start_date, end_date)
        data = {
            'project_status': project_data,
            'priority_data': priority_data
        }
    else:
        return jsonify({'error': 'Invalid report type'}), 400
    
    return jsonify(data)